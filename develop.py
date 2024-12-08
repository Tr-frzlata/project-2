import pandas as pd
from openpyxl import load_workbook


def find_id_column(df):
    # Ищем столбец, содержащий 'id' (без учета регистра)
    id_columns = df.columns[df.columns.str.lower() == 'id']
    if len(id_columns) > 0:
        return id_columns[0]
    return None


def update_ids(file_paths):
    last_id = None
    
    # Загрузка файла
    for i, file_path in enumerate(file_paths):
        df = pd.read_excel(file_path)

       
        # Сохранение изменений в файл
        book = load_workbook(file_path)
        sheet_name = book.sheetnames[0]

# Пути к файлам
file_paths = ['file1.xlsx', 'file2.xlsx', 'file3.xlsx']

# Вызов функции
update_ids(file_paths)

print("Обновление ID завершено во всех файлах.")
